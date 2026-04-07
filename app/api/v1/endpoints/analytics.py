"""
Analytics & Reporting endpoints.

GET  /analytics/overview          — dashboard KPI summary
GET  /analytics/compliance-score  — compliance gap analysis with AI recommendations
GET  /analytics/risk-distribution — knowledge entries by risk level
GET  /analytics/document-types    — breakdown by document type
POST /analytics/export            — generate PDF or Excel report (async, returns report ID)
GET  /analytics/reports/{id}      — poll report status + download URL
"""

import json
from datetime import datetime, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_current_active_user, require_role
from app.core.database import get_db
from app.models.compliance_rule import RuleCategory
from app.models.document import Document, DocumentStatus, DocumentType
from app.models.knowledge_entry import KnowledgeEntry, KnowledgeType, RiskLevel
from app.models.notification import Notification
from app.models.report import Report, ReportFormat, ReportStatus, ReportType
from app.models.user import User, UserRole
from app.services.compliance_service import compute_compliance_gaps, get_rules_for_company
from app.services.report_service import generate_excel_report, generate_pdf_report
import io

router = APIRouter()


# ---------------------------------------------------------------------------
# Overview
# ---------------------------------------------------------------------------

@router.get("/overview")
async def get_overview(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    if not current_user.company_id:
        return {
            "documents": {"total": 0, "processed": 0, "processing": 0, "uploaded": 0, "failed": 0, "processing_rate_pct": 0},
            "knowledge_entries": {"total": 0, "by_type": {}},
            "alerts": {"unread_notifications": 0, "upcoming_deadlines_30d": 0, "critical_risks": 0},
            "generated_at": datetime.utcnow().isoformat(),
        }

    cid = current_user.company_id

    # Document counts
    doc_counts = await db.execute(
        select(Document.status, func.count(Document.id))
        .where(Document.company_id == cid)
        .group_by(Document.status)
    )
    doc_stats = {str(s): c for s, c in doc_counts.all()}

    # Knowledge entry counts
    ke_counts = await db.execute(
        select(KnowledgeEntry.knowledge_type, func.count(KnowledgeEntry.id))
        .where(and_(KnowledgeEntry.company_id == cid, KnowledgeEntry.is_active.is_(True)))
        .group_by(KnowledgeEntry.knowledge_type)
    )
    ke_stats = {str(t): c for t, c in ke_counts.all()}

    # Unread notifications
    unread = await db.execute(
        select(func.count(Notification.id))
        .where(and_(Notification.user_id == current_user.id, Notification.is_read.is_(False)))
    )
    unread_count = unread.scalar() or 0

    # Upcoming deadlines within 30 days
    now = datetime.utcnow()
    upcoming = await db.execute(
        select(func.count(KnowledgeEntry.id))
        .where(
            and_(
                KnowledgeEntry.company_id == cid,
                KnowledgeEntry.is_active.is_(True),
                KnowledgeEntry.deadline.isnot(None),
                KnowledgeEntry.deadline >= now,
                KnowledgeEntry.deadline <= now + timedelta(days=30),
            )
        )
    )
    upcoming_deadlines = upcoming.scalar() or 0

    # Critical risks
    critical = await db.execute(
        select(func.count(KnowledgeEntry.id))
        .where(
            and_(
                KnowledgeEntry.company_id == cid,
                KnowledgeEntry.is_active.is_(True),
                KnowledgeEntry.risk_level == RiskLevel.CRITICAL,
            )
        )
    )
    critical_count = critical.scalar() or 0

    total_docs = sum(doc_stats.values())
    processed_docs = doc_stats.get("processed", 0)
    processing_rate = round((processed_docs / total_docs * 100) if total_docs else 0, 1)

    return {
        "documents": {
            "total": total_docs,
            "processed": processed_docs,
            "processing": doc_stats.get("processing", 0),
            "uploaded": doc_stats.get("uploaded", 0),
            "failed": doc_stats.get("failed", 0),
            "processing_rate_pct": processing_rate,
        },
        "knowledge_entries": {
            "total": sum(ke_stats.values()),
            "by_type": ke_stats,
        },
        "alerts": {
            "unread_notifications": unread_count,
            "upcoming_deadlines_30d": upcoming_deadlines,
            "critical_risks": critical_count,
        },
        "generated_at": datetime.utcnow().isoformat(),
    }


# ---------------------------------------------------------------------------
# Compliance score + gap analysis
# ---------------------------------------------------------------------------

@router.get("/compliance-score")
async def get_compliance_score(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    category: Optional[str] = Query(None, description="Filter by rule category"),
):
    if not current_user.company_id:
        return {
            "compliance_score": 0, "country": None, "total_rules": 0,
            "covered_rules": 0, "coverage_percentage": 0, "gap_rules": [],
        }

    # Load company
    from app.models.company import Company
    company_result = await db.execute(
        select(Company).where(Company.id == current_user.company_id)
    )
    company = company_result.scalar_one_or_none()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    # Load all knowledge titles for matching
    ke_result = await db.execute(
        select(KnowledgeEntry.title, KnowledgeEntry.content)
        .where(and_(KnowledgeEntry.company_id == company.id, KnowledgeEntry.is_active.is_(True)))
    )
    knowledge_texts = [f"{r.title} {r.content}" for r in ke_result.all()]

    gap_report = await compute_compliance_gaps(
        db=db,
        company_id=str(company.id),
        country_code=company.country or "US",
        industry=company.industry,
        knowledge_titles=knowledge_texts,
    )

    # Update stored compliance score
    company.compliance_score = gap_report["coverage_percentage"]
    await db.commit()

    return {
        "compliance_score": gap_report["coverage_percentage"],
        "country": company.country,
        **gap_report,
    }


# ---------------------------------------------------------------------------
# Time-series document & query activity
# ---------------------------------------------------------------------------

@router.get("/activity")
async def get_activity(
    days: int = Query(14, ge=7, le=90, description="Number of days to look back"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Daily counts of documents uploaded, processed, and knowledge entries created."""
    if not current_user.company_id:
        return {"days": days, "series": []}

    cid = current_user.company_id
    now = datetime.utcnow()
    start = now - timedelta(days=days)

    # Daily document uploads
    upload_sql = text("""
        SELECT DATE(created_at) AS day, COUNT(*) AS cnt
        FROM documents
        WHERE company_id = :cid AND created_at >= :start
        GROUP BY DATE(created_at)
        ORDER BY day
    """)
    upload_rows = (await db.execute(upload_sql, {"cid": str(cid), "start": start})).fetchall()

    # Daily documents that moved to processed
    processed_sql = text("""
        SELECT DATE(updated_at) AS day, COUNT(*) AS cnt
        FROM documents
        WHERE company_id = :cid AND status = 'processed' AND updated_at >= :start
        GROUP BY DATE(updated_at)
        ORDER BY day
    """)
    processed_rows = (await db.execute(processed_sql, {"cid": str(cid), "start": start})).fetchall()

    # Daily knowledge entries created (proxy for AI query volume / extraction activity)
    ke_sql = text("""
        SELECT DATE(created_at) AS day, COUNT(*) AS cnt
        FROM knowledge_entries
        WHERE company_id = :cid AND created_at >= :start
        GROUP BY DATE(created_at)
        ORDER BY day
    """)
    ke_rows = (await db.execute(ke_sql, {"cid": str(cid), "start": start})).fetchall()

    # Build a day-keyed dict for each series
    uploads_by_day    = {str(r.day): r.cnt for r in upload_rows}
    processed_by_day  = {str(r.day): r.cnt for r in processed_rows}
    entries_by_day    = {str(r.day): r.cnt for r in ke_rows}

    # Generate every day in the range
    series = []
    for i in range(days):
        day = (start + timedelta(days=i + 1)).date()
        day_str = str(day)
        series.append({
            "date":      day_str,
            "label":     day.strftime("%b %d"),
            "uploaded":  uploads_by_day.get(day_str, 0),
            "processed": processed_by_day.get(day_str, 0),
            "entries":   entries_by_day.get(day_str, 0),
        })

    return {"days": days, "series": series}


# ---------------------------------------------------------------------------
# Risk distribution
# ---------------------------------------------------------------------------

@router.get("/risk-distribution")
async def get_risk_distribution(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    if not current_user.company_id:
        return {"distribution": {}, "percentages": {}, "total": 0}

    result = await db.execute(
        select(KnowledgeEntry.risk_level, func.count(KnowledgeEntry.id))
        .where(
            and_(
                KnowledgeEntry.company_id == current_user.company_id,
                KnowledgeEntry.is_active.is_(True),
                KnowledgeEntry.risk_level.isnot(None),
            )
        )
        .group_by(KnowledgeEntry.risk_level)
    )
    distribution = {str(level): count for level, count in result.all()}
    total = sum(distribution.values()) or 1

    return {
        "distribution": distribution,
        "percentages": {k: round(v / total * 100, 1) for k, v in distribution.items()},
        "total": total,
    }


# ---------------------------------------------------------------------------
# Document type breakdown
# ---------------------------------------------------------------------------

@router.get("/document-types")
async def get_document_type_breakdown(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    if not current_user.company_id:
        return {"breakdown": {}}

    result = await db.execute(
        select(Document.document_type, func.count(Document.id))
        .where(Document.company_id == current_user.company_id)
        .group_by(Document.document_type)
    )
    return {"breakdown": {str(dtype): count for dtype, count in result.all()}}


# ---------------------------------------------------------------------------
# Report export (PDF / Excel)
# ---------------------------------------------------------------------------

@router.post("/export")
async def export_report(
    report_type: ReportType,
    report_format: ReportFormat,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Synchronously generate and stream a PDF or Excel report."""
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="User must belong to a company")

    from app.models.company import Company
    company_result = await db.execute(
        select(Company).where(Company.id == current_user.company_id)
    )
    company = company_result.scalar_one_or_none()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    # Build report data
    report_data = await _build_report_data(db, current_user.company_id, report_type)

    try:
        if report_format == ReportFormat.PDF:
            content = generate_pdf_report(report_data, company.name, report_type.value)
            media_type = "application/pdf"
            filename = f"{report_type.value}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        elif report_format == ReportFormat.EXCEL:
            content = generate_excel_report(report_data, company.name, report_type.value)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"{report_type.value}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        else:
            return report_data
    except ImportError as e:
        raise HTTPException(status_code=501, detail=str(e))

    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

async def _build_report_data(db: AsyncSession, company_id, report_type: ReportType):
    """Gather data for report generation."""
    data = {}

    # Summary block
    doc_count = await db.execute(
        select(func.count(Document.id)).where(Document.company_id == company_id)
    )
    ke_count = await db.execute(
        select(func.count(KnowledgeEntry.id))
        .where(and_(KnowledgeEntry.company_id == company_id, KnowledgeEntry.is_active.is_(True)))
    )
    data["summary"] = {
        "total_documents": doc_count.scalar() or 0,
        "total_knowledge_entries": ke_count.scalar() or 0,
        "report_type": report_type.value,
        "generated_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
    }

    if report_type in (ReportType.DOCUMENT_SUMMARY, ReportType.FULL_DASHBOARD):
        docs_result = await db.execute(
            select(
                Document.original_filename,
                Document.document_type,
                Document.status,
                Document.file_size,
                Document.created_at,
            )
            .where(Document.company_id == company_id)
            .order_by(Document.created_at.desc())
            .limit(500)
        )
        data["documents"] = [
            {
                "filename": r.original_filename,
                "type": str(r.document_type),
                "status": str(r.status),
                "size_kb": round((r.file_size or 0) / 1024, 1),
                "uploaded": r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
            }
            for r in docs_result.all()
        ]

    if report_type in (ReportType.RISK_DISTRIBUTION, ReportType.FULL_DASHBOARD, ReportType.KNOWLEDGE_AUDIT):
        ke_result = await db.execute(
            select(
                KnowledgeEntry.title,
                KnowledgeEntry.knowledge_type,
                KnowledgeEntry.risk_level,
                KnowledgeEntry.deadline,
                KnowledgeEntry.created_at,
            )
            .where(and_(KnowledgeEntry.company_id == company_id, KnowledgeEntry.is_active.is_(True)))
            .order_by(KnowledgeEntry.risk_level.desc())
            .limit(1000)
        )
        data["knowledge_entries"] = [
            {
                "title": r.title,
                "type": str(r.knowledge_type),
                "risk_level": str(r.risk_level),
                "deadline": r.deadline.strftime("%Y-%m-%d") if r.deadline else "N/A",
                "created": r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
            }
            for r in ke_result.all()
        ]

    return data


# ---------------------------------------------------------------------------
# Knowledge entries export (CSV / Excel)
# ---------------------------------------------------------------------------

@router.get(
    "/export-knowledge",
    summary="Export knowledge entries as CSV or Excel",
    description=(
        "Download all extracted knowledge entries (obligations, deadlines, risks, metrics) "
        "for the company as a CSV or Excel spreadsheet. "
        "Query param `format` accepts `csv` (default) or `excel`."
    ),
)
async def export_knowledge(
    format: str = "csv",
    risk_level: Optional[str] = None,
    knowledge_type: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    if not current_user.company_id:
        return StreamingResponse(iter([b""]), media_type="text/csv")

    filters = [
        KnowledgeEntry.company_id == current_user.company_id,
        KnowledgeEntry.is_active.is_(True),
    ]
    if risk_level:
        filters.append(KnowledgeEntry.risk_level == risk_level)
    if knowledge_type:
        filters.append(KnowledgeEntry.knowledge_type == knowledge_type)

    result = await db.execute(
        select(KnowledgeEntry).where(and_(*filters)).order_by(KnowledgeEntry.risk_level.desc())
    )
    entries = result.scalars().all()

    rows = [
        {
            "ID": str(e.id),
            "Title": e.title,
            "Type": str(e.knowledge_type),
            "Risk Level": str(e.risk_level) if e.risk_level else "",
            "Deadline": e.deadline.strftime("%Y-%m-%d") if e.deadline else "",
            "Content": e.content[:500] if e.content else "",
            "Tags": ", ".join(e.tags or []),
            "Created": e.created_at.strftime("%Y-%m-%d") if e.created_at else "",
        }
        for e in entries
    ]

    if format == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise HTTPException(status_code=501, detail="openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Knowledge Entries"
        headers = list(rows[0].keys()) if rows else ["ID", "Title", "Type", "Risk Level", "Deadline", "Content", "Tags", "Created"]
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for row_idx, row in enumerate(rows, 2):
            for col_idx, (_, val) in enumerate(row.items(), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"knowledge_entries_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # Default: CSV
    import csv
    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    csv_bytes = buf.getvalue().encode()
    filename = f"knowledge_entries_{datetime.utcnow().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Scheduled reports — CRUD
# ---------------------------------------------------------------------------

from pydantic import BaseModel as _BaseModel


class ScheduleReportRequest(_BaseModel):
    report_type: ReportType
    report_format: ReportFormat
    schedule: str           # cron expression, e.g. "0 8 * * 1"
    schedule_email: str     # comma-separated email recipients


@router.post(
    "/schedule",
    summary="Schedule a recurring report",
    description=(
        "Creates a scheduled report that will be auto-generated and emailed on a cron schedule. "
        "Example cron: `0 8 * * 1` = every Monday at 8am UTC. "
        "The report is emailed to `schedule_email`."
    ),
    status_code=201,
)
async def schedule_report(
    payload: ScheduleReportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="User must belong to a company")

    from app.models.report import Report, ReportStatus
    try:
        from croniter import croniter
        cron = croniter(payload.schedule)
        next_run = cron.get_next(datetime)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid cron expression. Example: '0 8 * * 1' (Mon 8am UTC)")

    report = Report(
        company_id=current_user.company_id,
        requested_by=current_user.id,
        report_type=payload.report_type,
        report_format=payload.report_format,
        status=ReportStatus.PENDING,
        schedule=payload.schedule,
        schedule_email=payload.schedule_email,
        next_run_at=next_run,
    )
    db.add(report)
    await db.commit()
    await db.refresh(report)
    return {
        "id": str(report.id),
        "report_type": report.report_type.value,
        "report_format": report.report_format.value,
        "schedule": report.schedule,
        "schedule_email": report.schedule_email,
        "next_run_at": report.next_run_at,
        "status": "scheduled",
    }


@router.get(
    "/schedules",
    summary="List scheduled reports",
    description="Returns all scheduled recurring reports for the current company.",
)
async def list_scheduled_reports(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    if not current_user.company_id:
        return []
    from app.models.report import Report
    result = await db.execute(
        select(Report).where(
            Report.company_id == current_user.company_id,
            Report.schedule.isnot(None),
        ).order_by(Report.created_at.desc())
    )
    reports = result.scalars().all()
    return [
        {
            "id": str(r.id),
            "report_type": r.report_type.value,
            "report_format": r.report_format.value,
            "schedule": r.schedule,
            "schedule_email": r.schedule_email,
            "next_run_at": r.next_run_at,
            "last_sent_at": r.last_sent_at,
        }
        for r in reports
    ]


@router.delete(
    "/schedules/{report_id}",
    summary="Delete a scheduled report",
    description="Removes a scheduled report so it will no longer be auto-generated.",
)
async def delete_scheduled_report(
    report_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    from app.models.report import Report
    result = await db.execute(
        select(Report).where(
            Report.id == report_id,
            Report.company_id == current_user.company_id,
        )
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="Scheduled report not found")
    await db.delete(report)
    await db.commit()
    return {"status": "deleted"}

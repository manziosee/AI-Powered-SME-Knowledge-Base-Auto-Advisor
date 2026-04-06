import logging
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook
from io import BytesIO
from typing import Optional

logger = logging.getLogger(__name__)

# Minimum characters from pypdf before we consider the PDF image-based
_OCR_FALLBACK_THRESHOLD = 100


async def _ocr_pdf(file_content: bytes) -> str:
    """OCR fallback for scanned/image-based PDFs using pytesseract or pdf2image."""
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
    except ImportError:
        logger.warning(
            "OCR libraries not installed. Install with: pip install pdf2image pytesseract pillow. "
            "Also ensure Tesseract-OCR is installed on the system."
        )
        return ""

    try:
        images = convert_from_bytes(file_content, dpi=200)
        text = ""
        for img in images:
            text += pytesseract.image_to_string(img) + "\n"
        return text.strip()
    except Exception as exc:
        logger.warning("OCR failed: %s", exc)
        return ""


async def extract_text_from_pdf(file_content: bytes) -> str:
    try:
        pdf = PdfReader(BytesIO(file_content))
        text = ""
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        text = text.strip()

        # If pypdf returned very little text, this is likely a scanned PDF — use OCR
        if len(text) < _OCR_FALLBACK_THRESHOLD:
            logger.info("pypdf extracted <100 chars — attempting OCR fallback")
            ocr_text = await _ocr_pdf(file_content)
            if len(ocr_text) > len(text):
                return ocr_text

        return text
    except Exception as exc:
        logger.warning("pypdf failed (%s) — attempting OCR fallback", exc)
        return await _ocr_pdf(file_content)


async def extract_text_from_docx(file_content: bytes) -> str:
    try:
        doc = DocxDocument(BytesIO(file_content))
        text = "\n".join([para.text for para in doc.paragraphs])
        return text.strip()
    except Exception:
        return ""


async def extract_text_from_xlsx(file_content: bytes) -> str:
    try:
        wb = load_workbook(BytesIO(file_content), read_only=True)
        text = ""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                text += " ".join([str(cell) for cell in row if cell]) + "\n"
        return text.strip()
    except Exception:
        return ""


async def extract_text(file_content: bytes, mime_type: str) -> Optional[str]:
    if mime_type == "application/pdf":
        return await extract_text_from_pdf(file_content)
    elif mime_type in ["application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/msword"]:
        return await extract_text_from_docx(file_content)
    elif mime_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
        return await extract_text_from_xlsx(file_content)
    elif mime_type.startswith("text/"):
        return file_content.decode('utf-8', errors='ignore')
    return None

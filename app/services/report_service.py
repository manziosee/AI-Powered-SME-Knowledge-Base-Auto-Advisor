import io
import json
from datetime import datetime
from typing import Any, Dict, List

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _header_fill():
    return PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")


def _alt_fill():
    return PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def _thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

def generate_pdf_report(report_data: Dict[str, Any], company_name: str, report_type: str) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is required for PDF export. Install it with: pip install reportlab")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "SectionHeader",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=colors.HexColor("#2E75B6"),
        spaceBefore=14,
        spaceAfter=6,
    )

    elements = []

    # Title block
    elements.append(Paragraph(f"{company_name} — {report_type.replace('_', ' ').title()}", title_style))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1F4E79")))
    elements.append(Spacer(1, 0.4 * cm))

    # Summary KPIs
    if "summary" in report_data:
        elements.append(Paragraph("Summary", section_style))
        summary = report_data["summary"]
        kpi_data = [["Metric", "Value"]]
        for k, v in summary.items():
            kpi_data.append([k.replace("_", " ").title(), str(v)])

        table = Table(kpi_data, colWidths=[9 * cm, 7 * cm])
        table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF2FA")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ])
        )
        elements.append(table)
        elements.append(Spacer(1, 0.4 * cm))

    # Detail sections
    for section_key, section_value in report_data.items():
        if section_key == "summary":
            continue
        if isinstance(section_value, list) and section_value:
            elements.append(Paragraph(section_key.replace("_", " ").title(), section_style))
            if isinstance(section_value[0], dict):
                headers = list(section_value[0].keys())
                rows = [[str(row.get(h, "")) for h in headers] for row in section_value]
                col_count = len(headers)
                col_width = 17 * cm / col_count
                tdata = [[h.replace("_", " ").title() for h in headers]] + rows
                table = Table(tdata, colWidths=[col_width] * col_count, repeatRows=1)
                table.setStyle(
                    TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF2FA")]),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#BFBFBF")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 5),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("WORDWRAP", (0, 0), (-1, -1), True),
                    ])
                )
                elements.append(table)
            elements.append(Spacer(1, 0.3 * cm))

    doc.build(elements)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def generate_excel_report(report_data: Dict[str, Any], company_name: str, report_type: str) -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1F4E79")

    # ------------------------------------------------------------------
    # Summary sheet
    # ------------------------------------------------------------------
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    ws_summary["A1"] = f"{company_name} — {report_type.replace('_', ' ').title()}"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws_summary["A2"].font = Font(italic=True, color="808080")

    if "summary" in report_data:
        row = 4
        ws_summary.cell(row, 1, "Metric").font = Font(bold=True, color="FFFFFF")
        ws_summary.cell(row, 2, "Value").font = Font(bold=True, color="FFFFFF")
        for col in [1, 2]:
            ws_summary.cell(row, col).fill = _header_fill()
            ws_summary.cell(row, col).alignment = Alignment(horizontal="center")

        for k, v in report_data["summary"].items():
            row += 1
            ws_summary.cell(row, 1, k.replace("_", " ").title()).border = _thin_border()
            ws_summary.cell(row, 2, str(v)).border = _thin_border()
            if row % 2 == 0:
                for col in [1, 2]:
                    ws_summary.cell(row, col).fill = _alt_fill()

    # ------------------------------------------------------------------
    # Detail sheets per section
    # ------------------------------------------------------------------
    for section_key, section_value in report_data.items():
        if section_key == "summary":
            continue
        if isinstance(section_value, list) and section_value and isinstance(section_value[0], dict):
            sheet_name = section_key.replace("_", " ").title()[:31]
            ws = wb.create_sheet(sheet_name)
            headers = list(section_value[0].keys())

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(1, col_idx, header.replace("_", " ").title())
                cell.font = header_font
                cell.fill = _header_fill()
                cell.alignment = Alignment(horizontal="center")
                cell.border = _thin_border()
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

            for row_idx, item in enumerate(section_value, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row_idx, col_idx, str(item.get(header, "")))
                    cell.border = _thin_border()
                    if row_idx % 2 == 0:
                        cell.fill = _alt_fill()

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
